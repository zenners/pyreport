from flask import Flask, request, jsonify, send_file
import json
import requests
import pandas as pd
import openpyxl
import flask_excel as excel
from io import BytesIO
import os


import smtplib,ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

app=Flask(__name__)
excel.init_excel(app)
port =  5001
#port = int(os.getenv("PORT"))

def send_mail(send_from,send_to,subject,text,filename,server,port,username='',password='',isTls=True):
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = send_to
    msg['Date'] = formatdate(localtime = True)
    msg['Subject'] = subject
    msg.attach(MIMEText(text))

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(filename, "rb").read())
    encoders.encode_base64(part)
    the_file = 'attachment; filename="{}"'.format(filename)
    part.add_header('Content-Disposition', the_file)
    msg.attach(part)

    #context = ssl.SSLContext(ssl.PROTOCOL_SSLv3)
    #SSL connection only working on Python 3+
    smtp = smtplib.SMTP(server, port)
    if isTls:
        smtp.starttls()
    smtp.login(username,password)
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.quit()

@app.route("/", methods=['GET'])
def index():
    return 'Hello World! I am running on port ' + str(port)

@app.route("/unappliedbalances", methods=['GET'])
def get_uabalances():
    output = BytesIO()

    url = "https://api360.zennerslab.com/Service1.svc/accountDueReportJSON"
    r = requests.post(url)
    data = r.json()
    greater_than_zero = list(filter(lambda x: x['unappliedBalance'] > 0, data['accountDueReportJSONResult']))

    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    headers = ["Loan Account Number", "Customer Name", "Mobile No.", "Amount Due", "Due Date", "Unapplied Balance"]
    df = pd.DataFrame(data['accountDueReportJSONResult'])

    # return jsonify(greater_than_zero)

    df = df[["loanAccountNo", "name", "mobileNo", "amountDue", "dueDate", "unappliedBalance"]]
    df.to_excel(writer, startrow=5, merge_cells=False, index=False, sheet_name = "Sheet_1", header=headers)

    workbook = writer.book
    merge_format = workbook.add_format({'align': 'center'})
    #xldate_header = "{} to {}".format(dateStart, dateEnd)
    xldate_header = "Today"


    worksheet = writer.sheets["Sheet_1"]
    worksheet.merge_range('B1:E1', 'RADIOWEALTH FINANCE COMPANY, INC.', merge_format)
    worksheet.merge_range('B2:E2', 'RFC360 Kwikredit', merge_format)
    worksheet.merge_range('B3:E3', 'Unapplied Balances Report', merge_format)
    worksheet.merge_range('B4:E4', xldate_header , merge_format)

    #the writer has done its job
    writer.close()

    #go back to the beginning of the stream
    output.seek(0)
    print('sending spreadsheet')
    filename = "Unapplied Balance.xlsx"
    return send_file(output, attachment_filename=filename, as_attachment=True)


@app.route("/dccr", methods=['GET'])
def get_data():
    output = BytesIO()
    dateStart = request.args.get('startDate')
    dateEnd = request.args.get('endDate')
    payload = {'startDate': dateStart, 'endDate': dateEnd}
    url = "https://api360.zennerslab.com/Service1.svc/DCCRjson"
    r = requests.post(url, json=payload)
    data_json = r.json()

    # pandas to excel
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    headers = ["Loan Account Number", "Customer Name", "Mobile No.", "OR Number", "OR Date", "Net Cash", "Payment Source"]
    df = pd.DataFrame(data_json['DCCRjsonResult'])
    df = df[['loanAccountNo', 'customerName','mobileNo', 'orNo' ,"postedDate","amountApplied", "paymentSource"]]
    df.to_excel(writer, startrow=5, merge_cells=False, index=False, sheet_name = "Sheet_1", header=headers)

    workbook = writer.book
    merge_format = workbook.add_format({'align': 'center'})
    xldate_header = "{} to {}".format(dateStart, dateEnd)


    worksheet = writer.sheets["Sheet_1"]
    worksheet.merge_range('C1:F1', 'RADIOWEALTH FINANCE COMPANY, INC.', merge_format)
    worksheet.merge_range('C2:F2', 'RFC360 Kwikredit', merge_format)
    worksheet.merge_range('C3:F3', 'Daily Cash Collection Report', merge_format)
    worksheet.merge_range('C4:F4', xldate_header , merge_format)

    #the writer has done its job
    writer.close()

    #go back to the beginning of the stream
    output.seek(0)
    print('sending spreadsheet')
    filename = "DCCR {}-{}.xlsx".format(dateStart, dateEnd)
    return send_file(output, attachment_filename=filename, as_attachment=True)

@app.route("/dccr2", methods=['GET'])
def get_data2():

    output = 'test.xlsx'
    dateStart = request.args.get('startDate')
    dateEnd = request.args.get('endDate')
    filename = "DCCR {}-{}.xlsx".format(dateStart, dateEnd)

    payload = {'startDate': dateStart, 'endDate': dateEnd}
    url = "https://api360.zennerslab.com/Service1.svc/DCCRjson"
    r = requests.post(url, json=payload)
    data_json = r.json()

    # pandas to excel
    writer = pd.ExcelWriter(filename, engine='xlsxwriter')
    headers = ["Loan Account Number", "Customer Name", "Mobile No.", "OR Number", "OR Date", "Net Cash", "Payment Source"]
    df = pd.DataFrame(data_json['DCCRjsonResult'])
    df = df[['loanAccountNo', 'customerName','mobileno','orNo' ,"postedDate","amountApplied", "paymentSource"]]
    df.to_excel(writer, startrow=5, merge_cells=False, index=False, sheet_name = "Sheet_1", header=headers)

    workbook = writer.book
    merge_format = workbook.add_format({'align': 'center'})
    xldate_header = "{} to {}".format(dateStart, dateEnd)


    worksheet = writer.sheets["Sheet_1"]
    worksheet.merge_range('C1:F1', 'RADIOWEALTH FINANCE COMPANY, INC.', merge_format)
    worksheet.merge_range('C2:F2', 'RFC360 Kwikredit', merge_format)
    worksheet.merge_range('C3:F3', 'Daily Cash Collection Report', merge_format)
    worksheet.merge_range('C4:F4', xldate_header , merge_format)

    #the writer has done its job
    writer.save()

    #go back to the beginning of the stream
    # output.seek(0)
    print('sending spreadsheet')
    send_mail("cu.michaels@gmail.com", "jantzen@thegentlemanproject.com", "hello", "helloworld", filename, 'smtp.gmail.com', '587', 'cu.michaels@gmail.com', 'jantzen216' )
    return 'ok'
    #return send_file(output, attachment_filename=filename, as_attachment=True)


@app.route("/monthlyincome", methods=['GET'])
def get_monthly():
    output = BytesIO()
    date = request.args.get('date')

    payload = {'date': date}
    url = "https://api360.zennerslab.com/Service1.svc/monthlyIncomeReportJs"
    r = requests.post(url, json=payload)
    data_json = r.json()

    #return r.text
    # pandas to excel
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    headers = ["Application ID", "Loan Account Number", "Customer Name", "Payment Amount", "Penalty Paid", "Interest Paid", "Principal Paid"]
    df = pd.DataFrame(data_json['monthlyIncomeReportJsResult'])
    df = df[['appId','loanAccountno', 'name', "paymentAmount", "penaltyPaid","interestPaid", "principalPaid"]]
    df.to_excel(writer, startrow=5, merge_cells=False, index=False, sheet_name = "Sheet_1", header=headers)

    workbook = writer.book
    merge_format = workbook.add_format({'align': 'center'})
    xldate_header = "{}".format(date)


    worksheet = writer.sheets["Sheet_1"]
    worksheet.merge_range('C1:F1', 'RADIOWEALTH FINANCE COMPANY, INC.', merge_format)
    worksheet.merge_range('C2:F2', 'RFC360 Kwikredit', merge_format)
    worksheet.merge_range('C3:F3', 'Monthly Income Report', merge_format)
    worksheet.merge_range('C4:F4', xldate_header , merge_format)
    #
    # #the writer has done its job
    writer.close()
    #
    # #go back to the beginning of the stream
    output.seek(0)
    print('sending spreadsheet')
    filename = "Monthly Income {}.xlsx".format(date)
    return send_file(output, attachment_filename=filename, as_attachment=True)


@app.route("/booking", methods=['GET'])
def get_booking():
    output = BytesIO()
    dateStart = request.args.get('startDate')
    dateEnd = request.args.get('endDate')
    payload = {'startDate': dateStart, 'endDate': dateEnd}
    url = "https://api360.zennerslab.com/Service1.svc/bookingReportJs"
    r = requests.post(url, json=payload)
    data_json = r.json()
    #return r.text
    # pandas to excel
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    headers = ["Application ID", "Loan Account Number", "Customer Name", "Subproduct", "PNV", "MLV", "Finance Fee", "GCLI", "Handling Fee", "Term", "Rate", "Monthly Amortization", "Booking Date", "Approval Date", "Application Date", "Branch"]
    df = pd.DataFrame(data_json['bookingReportJsResult'])
    df = df[['loanId', 'loanAccountNo','customerName' ,"subProduct","PNV", "principal", "interest", "insurance", "handlingFee", "term", "actualRate", "monthlyAmount", "forreleasingdate", 'approvalDate', 'applicationDate', 'branch']]
    df.to_excel(writer, startrow=5, merge_cells=False, index=False, sheet_name = "Sheet_1", header=headers)

    workbook = writer.book
    merge_format = workbook.add_format({'align': 'center'})
    xldate_header = "{} to {}".format(dateStart, dateEnd)


    worksheet = writer.sheets["Sheet_1"]
    worksheet.merge_range('C1:F1', 'RADIOWEALTH FINANCE COMPANY, INC.', merge_format)
    worksheet.merge_range('C2:F2', 'RFC360 Kwikredit', merge_format)
    worksheet.merge_range('C3:F3', 'Booking Report  ', merge_format)
    worksheet.merge_range('C4:F4', xldate_header , merge_format)
    #
    # #the writer has done its job
    writer.close()
    #
    # #go back to the beginning of the stream
    output.seek(0)
    print('sending spreadsheet')
    filename = "Booking report {}-{}.xlsx".format(dateStart, dateEnd)
    return send_file(output, attachment_filename=filename, as_attachment=True)


@app.route("/incentive", methods=['GET'])
def get_incentive():
    output = BytesIO()
    dateStart = request.args.get('startDate')
    dateEnd = request.args.get('endDate')
    payload = {'startDate': dateStart, 'endDate': dateEnd}
    url = "https://api360.zennerslab.com/Service1.svc/generateincentiveReportJSON"
    r = requests.post(url, json=payload)
    data_json = r.json()
    #return r.text
    # pandas to excel
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    headers = ["Application ID", "Full Name", "Referral Type", "SA", "Branch", "Term", "Finance Amount", "PNV", "Monthly Installment", "Promodiser Name"]
    df = pd.DataFrame(data_json['generateincentiveReportJSONResult'])
    df = df[['loanId', 'borrowerName','refferalType' ,"SA","dealerName", "term", "totalAmount", "PNV", "monthlyAmount", "agentName"]]
    df.to_excel(writer, startrow=5, merge_cells=False, index=False, sheet_name = "Sheet_1", header=headers)

    workbook = writer.book
    merge_format = workbook.add_format({'align': 'center'})
    xldate_header = "{} to {}".format(dateStart, dateEnd)


    worksheet = writer.sheets["Sheet_1"]
    worksheet.merge_range('C1:F1', 'RADIOWEALTH FINANCE COMPANY, INC.', merge_format)
    worksheet.merge_range('C2:F2', 'RFC360 Kwikredit', merge_format)
    worksheet.merge_range('C3:F3', 'Request for Merchandiser Commission  ', merge_format)
    worksheet.merge_range('C4:F4', xldate_header , merge_format)
    #
    # #the writer has done its job
    writer.close()
    #
    # #go back to the beginning of the stream
    output.seek(0)
    print('sending spreadsheet')
    filename = "Merchandiser Commission report {}-{}.xlsx".format(dateStart, dateEnd)
    return send_file(output, attachment_filename=filename, as_attachment=True)


@app.route("/mature", methods=['GET'])
def get_mature():
    output = BytesIO()
    date = request.args.get('date')
    payload = {'date': date}
    url = "https://api360.zennerslab.com/Service1.svc/maturedLoanReport"
    r = requests.post(url, json=payload)
    data_json = r.json()
    #return r.text
    # pandas to excel
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    headers = ["Application ID", "Loan Account Number", "Customer Name", "Mobile No.", "Term", "Last Due Date", "Last Payment", "No. of Unpaid Months", "Monthly Due", "Outstanding Balance"]
    df = pd.DataFrame(data_json['maturedLoanReportResult'])
    df = df[['loanId', 'loanAccountNo','fullName',"mobileno","term","lastDueDate","lastPayment", "unpaidMonths", "monthlydue", "outStandingBalance" ]]
    df.to_excel(writer, startrow=5, merge_cells=False, index=False, sheet_name = "Sheet_1", header=headers)

    workbook = writer.book
    merge_format = workbook.add_format({'align': 'center'})
    xldate_header = "As of {}".format(date)


    worksheet = writer.sheets["Sheet_1"]
    worksheet.merge_range('C1:F1', 'RADIOWEALTH FINANCE COMPANY, INC.', merge_format)
    worksheet.merge_range('C2:F2', 'RFC360 Kwikredit', merge_format)
    worksheet.merge_range('C3:F3', 'Matured Loans Report  ', merge_format)
    worksheet.merge_range('C4:F4', xldate_header , merge_format)
    #
    # #the writer has done its job
    writer.close()
    #
    # #go back to the beginning of the stream
    output.seek(0)
    print('sending spreadsheet')
    filename = "Matured Loans report {}.xlsx".format(date)
    return send_file(output, attachment_filename=filename, as_attachment=True)


@app.route("/duetoday", methods=['GET'])
def get_due():
    output = BytesIO()
    date = request.args.get('date')
    payload = {'date': date}
    url = "https://api360.zennerslab.com/Service1.svc/dueTodayReport"
    r = requests.post(url, json=payload)
    data_json = r.json()
    #return r.text
    # pandas to excel
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    headers = ["Application ID", "Loan Account Number", "Customer Name", "Mobile No.", "Loan Type", "Term", "Monthly Installment", "Monthly Due"] 
    df = pd.DataFrame(data_json['dueTodayReportResult'])
    df = df[['loanId', 'loanAccountNo','fullName',"mobileno","loanType", "term", "monthlyAmmortization", "monthlydue" ]]
    df.to_excel(writer, startrow=5, merge_cells=False, index=False, sheet_name = "Sheet_1", header=headers)

    workbook = writer.book
    merge_format = workbook.add_format({'align': 'center'})
    xldate_header = "As of {}".format(date)


    worksheet = writer.sheets["Sheet_1"]
    worksheet.merge_range('C1:F1', 'RADIOWEALTH FINANCE COMPANY, INC.', merge_format)
    worksheet.merge_range('C2:F2', 'RFC360 Kwikredit', merge_format)
    worksheet.merge_range('C3:F3', 'Due Today Report  ', merge_format)
    worksheet.merge_range('C4:F4', xldate_header , merge_format)
    #
    # #the writer has done its job
    writer.close()
    #
    # #go back to the beginning of the stream
    output.seek(0)
    print('sending spreadsheet')
    filename = "Due Today report {}.xlsx".format(date)
    return send_file(output, attachment_filename=filename, as_attachment=True)



if __name__ == "__main__":
    app.run(host='0.0.0.0', port=port)
